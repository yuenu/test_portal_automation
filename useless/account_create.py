from openpyxl import load_workbook

# 先利用CreateFile function創建跟example.xlsx檔同樣格式的檔案,再將Data寫入

# Create same type xlsl file
def CreateFile():
    for i in range(41, 61):
        wb = load_workbook('./AccountCreate/AccountExaple.xlsx')
        wb.save(f'./AccountCreate/Accounttest{i}.xlsx')


def WriteData():
    account = 'QA_DS'

    # 從會員數多少開始
    nums_start = 1

    # 一個Excel檔要產生多少個會員
    One_time_num = 200000

    for times in range(41, 61):
        save_path = f'./AccountCreate/accounttest{times}.xlsx'
        wb = load_workbook(save_path)
        ws = wb.active

        for index, num in enumerate(range(nums_start, nums_start + One_time_num), 1):
            # 對應Excel的格式寫入檔案
            ws[f'A{index+1}'] = 'QAfourmillionA'
            ws[f'B{index+1}'] = 'QAfourmillionB'
            ws[f'C{index+1}'] = 'QAfourmillionC'
            ws[f'D{index+1}'] = 'QAfourmillionD'
            ws[f'E{index+1}'] = account + str(num).zfill(10)
            ws[f'F{index+1}'] = 'a123456'
            ws[f'G{index+1}'] = 'a123456'
            
            ws[f'H{index+1}'] = '王小明'
            ws[f'I{index+1}'] = '1234567890'
            ws[f'J{index+1}'] = 'example1@qq.com'
            ws[f'K{index+1}'] = '1234567890'
            ws[f'L{index+1}'] = '0'
            ws[f'M{index+1}'] = '是'

            ws[f'N{index+1}'] = '男'
            ws[f'O{index+1}'] = '2012/12/12'
            ws[f'P{index+1}'] = 'exampleweixin1'
            ws[f'Q{index+1}'] = '建设银行'
            ws[f'R{index+1}'] = '山西省'
            ws[f'S{index+1}'] = '太原市'
            ws[f'T{index+1}'] = '99876543210'
            ws[f'U{index+1}'] = '帐号备注1'
        
        wb.save(save_path)
        nums_start =  nums_start + One_time_num

# CreateFile()
WriteData()